from selenium import webdriver
import time
import re
import datetime
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import os
# Set the path to the Chrome binary
chrome_options = webdriver.ChromeOptions()
chrome_options.binary_location = '/usr/bin/chromium'
#chrome_options.binary_location = 'C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe'
home_page='https://www.umgc.edu/login'
login_email='cnanakumo@student.umgc.edu'
login_pswd='Samhilda@2023'
class_code=708985 #adjust accordingly
topic_code=3860892 #adjust accordingly
filters='unread'
discussion_url=f'https://learn.umgc.edu/d2l/le/{class_code}/discussions/topics/{topic_code}/View?filters={filters}&groupFilterOption=0'
course_url=f'https://app.schoology.com/course/{class_code}/members'
# excel timeStamp
current_time = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')



# Set up the Chrome driver with the downloaded ChromeDriver executable

#= Navigate to the Schoology login page







def startPageNav(startPage):
    global currentPage
    wait = WebDriverWait(driver, 10)
    try:
        print("Going to startPage")
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.next')))
        for page in range(startPage-1): #changed startPage-1
            next_page = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.next')))
            if next_page:
                next_page.click()
                time.sleep(5)
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.next')))
        print(f" Now on page> {currentPage}")
    except (ValueError,TypeError) as e:
        print("Error going to start page",e)


              
def nextPage():
    wait = WebDriverWait(driver, 7)
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.next')))
        next_page = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.next')))
        if next_page:
                next_page.click()
                print("Clicked next.")
                time.sleep(5)
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.next')))
    except:
        print("No more next pages*")

def print_data_to_excel(names, emails,fileName):
    current_dir = os.getcwd()
    file_path = os.path.join(current_dir, fileName)
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add headers
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Email"
    

    # Print data to respective columns
    for i in range(len(names)):
        sheet.cell(row=i+2, column=1).value = names[i]
        sheet.cell(row=i+2, column=2).value = emails[i]
       


    # Save the workbook
    workbook.save(file_path)
    print(f"Data printed to Excel successfully at:{current_time}")


def getData():
    ids=set() # a unique set of ids
    print("Script starting***")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    wait = WebDriverWait(driver, 5)
    driver.get(home_page)
    time.sleep(3)
    
    
    # pageToStart=input("Enter startPage: ")
    # pages=input("Enter the number of pages to go through. ie pages=totalScholars/ScholarsPerPage : ")
    # pageToStart=int(pageToStart)
    # currentPage=pageToStart
    # pages=int(pages)
    #filename='courseCode_'+str(class_code)+'_startPage_'+str(pageToStart)+'_'+ current_time + '.xlsx'
    file_name='umgc_contacts.xlsx'
    
    # login
    login_page=driver.find_element(By.ID, 'button-a198d4e78c').get_attribute('href')
    driver.get(login_page)
    time.sleep(3)
    WebDriverWait(driver, 7).until(EC.presence_of_element_located((By.ID, 'idSIButton9')))
    email_field = driver.find_element(By.ID, 'i0116')
    email_field.send_keys(login_email)
    to_passwd_btn=driver.find_element(By.ID, 'idSIButton9')
    time.sleep(3)
    to_passwd_btn.click()
    time.sleep(3)
    password_field = driver.find_element(By.ID, 'i0118')
    password_field.send_keys(login_pswd)
    print("waiting for the delay")
    submit_btn = driver.find_element(By.CSS_SELECTOR, 'input#idSIButton9')
    submit_btn.click()
    #probably add code to check if certain tag exists before navigating
    print("login done")
    time.sleep(3)
    driver.get(discussion_url)
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'a.d2l-linkheading-link')))
    post_heading=driver.find_element(By.CLASS_NAME, 'd2l-linkheading-link') #gets first post heading link
    time.sleep(5)
    # post_heading.click() #clicks on first link
    test_url='https://learn.umgc.edu/d2l/le/708985/discussions/threads/28379277/View?searchText=testing+the+waters'
    driver.get(test_url)
    time.sleep(3)
   
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.d2l-user-profile-handle-normal')))
    user_avartars = driver.find_elements(By.CLASS_NAME,'d2l-user-profile-handle-normal')
    print(f"no of avatars: {len(user_avartars)}")
    for avartar in user_avartars:
        print(f"avartar id before click: {avartar.get_attribute('id')}")

    
    for avatar in reversed(user_avartars):
        # print(f"avartar id: {avatar.get_attribute('id')}")
        avatar.click() #maybe wait until certain elements of the avarter are visible before proceeding to ensure that the placeholder in the DOM
        time.sleep(7)
        # avatar.click()
        # time.sleep(5)
      
        #maybe click the avartar again to close it
    id_elements=driver.find_elements(By.CSS_SELECTOR, '.d2l-placeholder.d2l-placeholder-inner.d2l-placeholder-live')
    for id_element in id_elements:
        placeholder_id=id_element.get_attribute('id')
        if 'profilePlaceholder' in placeholder_id:
            ids.add(placeholder_id.replace('profilePlaceholder',''))
    print(f"ids: {ids}")
    names=[]
    emails=[]
    for id in ids:
        
        driver.execute_script(f"window.open('https://learn.umgc.edu/d2l/lms/email/integration/AdaptLegacyPopupData.d2l?ou={class_code}&p=0&ext=1&cb=d2l_2_0_637&singleUserId={id}', 'new_window');")
        window_handles = driver.window_handles
        # Switch to the second window
        new_window = window_handles[1]
        driver.switch_to.window(new_window)
        time.sleep(7)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.d2l-multiselect-choice')))
        element=driver.find_element(By.CSS_SELECTOR, '.d2l-multiselect-choice') #d2l-heading vui-heading-1
        contact=element.text
        print(contact)
        match = re.match(r'"(.*?)" <(.*?)>', contact)
        if match:
            names.append(match.group(1)) 
            emails.append(match.group(2).replace('<', '').replace('>', ''))
        else:
            return print("regex not working")

        # time.sleep(20)

        driver.close()

    # Switch to the first window
        first_window = window_handles[0]
        driver.switch_to.window(first_window)
    print(f"names: {names}")
    print(f"emails: {emails}")
    print_data_to_excel(names,emails,file_name)

        # driver.close()

    # time.sleep(3) 
 
    # Submit the login form
    # driver.find_element(By.ID, 'edit-submit').click()
    # driver.implicitly_wait(10)
    # driver.get(course_url)
    # if pageToStart==1:
    #     for page in range(pages):
    #         loadData(page)
    # else:
    #     startPageNav(page)
    #     time.sleep(5)
    #     for page in range(pages):
    #         loadData(page)      
    # print_data_to_excel(names,emails,phones,filename=filename)
    # Quit
    # driver.quit()


getData()




